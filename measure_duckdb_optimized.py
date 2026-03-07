#!/usr/bin/env python3
"""
DuckDB Excel Formula Evaluation Benchmark - OPTIMIZED VERSION

Pure SQL approach using bulk operations:
1. Load Excel directly into DuckDB (via pandas)
2. Detect formula patterns and convert to SQL
3. Execute single SQL query per formula column
4. Bulk export results

Target: 10-100x faster than per-cell approach
"""

import re
import duckdb
import pandas as pd
import psutil
import json
import sys
import time
import threading
import openpyxl
from pathlib import Path
from typing import Any, Dict, List, Tuple, Optional, Pattern


# ─────────────────────────────────────────────────────────────────
# 1. Formula Pattern Detection & SQL Generation
# ─────────────────────────────────────────────────────────────────

class FormulaPatternDetector:
    """Detect Excel formula patterns and generate equivalent SQL queries."""

    # Pattern: =A{i}+B{i} → column addition
    ADDITION_PATTERN = re.compile(r'^=([A-Z])\d+\+([A-Z])\d+$', re.I)

    # Pattern: =A{i}*2 (or any literal multiplier)
    MULTIPLICATION_PATTERN = re.compile(r'^=([A-Z])\d+\*(\d+(?:\.\d+)?)$', re.I)

    # Pattern: =Sheet1!A{i} → cross-sheet reference
    CROSS_SHEET_PATTERN = re.compile(r'^=([A-Za-z0-9_]+)!([A-Z])\d+$', re.I)

    # Pattern: =A{i}*B{i} → column multiplication
    COL_MULTIPLY_PATTERN = re.compile(r'^=([A-Z])\d+\*([A-Z])\d+$', re.I)

    # Pattern: =A{i}-B{i} → column subtraction
    SUBTRACTION_PATTERN = re.compile(r'^=([A-Z])\d+-([A-Z])\d+$', re.I)

    # Pattern: =A{i}/B{i} → column division
    DIVISION_PATTERN = re.compile(r'^=([A-Z])\d+/([A-Z])\d+$', re.I)

    @staticmethod
    def detect_pattern(formula: str) -> Optional[Tuple[str, Dict[str, str]]]:
        """
        Detect the formula pattern.
        Returns (pattern_type, params) or None if unknown pattern.
        """
        m = FormulaPatternDetector.ADDITION_PATTERN.match(formula)
        if m:
            return ('addition', {'col1': m.group(1).lower(), 'col2': m.group(2).lower()})

        m = FormulaPatternDetector.SUBTRACTION_PATTERN.match(formula)
        if m:
            return ('subtraction', {'col1': m.group(1).lower(), 'col2': m.group(2).lower()})

        m = FormulaPatternDetector.COL_MULTIPLY_PATTERN.match(formula)
        if m:
            return ('col_multiply', {'col1': m.group(1).lower(), 'col2': m.group(2).lower()})

        m = FormulaPatternDetector.DIVISION_PATTERN.match(formula)
        if m:
            return ('division', {'col1': m.group(1).lower(), 'col2': m.group(2).lower()})

        m = FormulaPatternDetector.MULTIPLICATION_PATTERN.match(formula)
        if m:
            return ('multiply_literal', {'col': m.group(1).lower(), 'literal': m.group(2)})

        m = FormulaPatternDetector.CROSS_SHEET_PATTERN.match(formula)
        if m:
            return ('cross_sheet', {
                'source_sheet': m.group(1).lower().replace(' ', '_'),
                'source_col': m.group(2).lower()
            })

        return None

    @staticmethod
    def generate_sql(
        pattern_type: str,
        params: Dict[str, str],
        target_sheet: str,
        target_col: str,
        existing_cols: List[str]
    ) -> Optional[str]:
        """Generate SQL query for the detected pattern."""
        # Build column list (exclude the target column if it exists)
        cols_to_select = [f'"{c}"' for c in existing_cols if c != target_col]

        if pattern_type == 'addition':
            col1 = params['col1_sanitized']
            col2 = params['col2_sanitized']
            cols_to_select.append(f'"{col1}" + "{col2}" AS "{target_col}"')
            return f'SELECT {", ".join(cols_to_select)} FROM {target_sheet}'

        elif pattern_type == 'subtraction':
            col1 = params['col1_sanitized']
            col2 = params['col2_sanitized']
            cols_to_select.append(f'"{col1}" - "{col2}" AS "{target_col}"')
            return f'SELECT {", ".join(cols_to_select)} FROM {target_sheet}'

        elif pattern_type == 'col_multiply':
            col1 = params['col1_sanitized']
            col2 = params['col2_sanitized']
            cols_to_select.append(f'"{col1}" * "{col2}" AS "{target_col}"')
            return f'SELECT {", ".join(cols_to_select)} FROM {target_sheet}'

        elif pattern_type == 'division':
            col1 = params['col1_sanitized']
            col2 = params['col2_sanitized']
            cols_to_select.append(f'"{col1}" / NULLIF("{col2}", 0) AS "{target_col}"')
            return f'SELECT {", ".join(cols_to_select)} FROM {target_sheet}'

        elif pattern_type == 'multiply_literal':
            col = params['col_sanitized']
            literal = params['literal']
            cols_to_select.append(f'"{col}" * {literal} AS "{target_col}"')
            return f'SELECT {", ".join(cols_to_select)} FROM {target_sheet}'

        elif pattern_type == 'cross_sheet':
            source_sheet = params['source_sheet']
            source_col = params['source_col_sanitized']
            # Build column list from target table (exclude the target column that will be computed)
            t1_cols = [f't1."{c}"' for c in existing_cols if c != target_col and c != '_row']
            # Add the row number and the computed column from source table
            t1_cols.insert(0, 't1._row')
            t1_cols.append(f't2."{source_col}" AS "{target_col}"')
            # Use row number for JOIN (assuming aligned rows)
            return f"""
                SELECT {", ".join(t1_cols)}
                FROM {target_sheet} t1
                JOIN {source_sheet} t2 ON t1._row = t2._row
            """

        return None


# ─────────────────────────────────────────────────────────────────
# 2. Excel Bulk Loading
# ─────────────────────────────────────────────────────────────────

def sanitize_column_name(name: str) -> str:
    """Sanitize column name for SQL use."""
    # Remove special chars, keep alphanumeric and underscore
    sanitized = re.sub(r'[^a-zA-Z0-9_]', '_', str(name))
    # Ensure it starts with a letter
    if sanitized and sanitized[0].isdigit():
        sanitized = 'col_' + sanitized
    return sanitized.lower() if sanitized else 'column'


def load_excel_to_duckdb(excel_path: Path, conn: duckdb.DuckDBPyConnection) -> Tuple[Dict[str, pd.DataFrame], Dict[str, Dict[str, str]], Dict[str, int]]:
    """
    Load Excel sheets into DuckDB using pandas (bulk read).
    Returns (dict of {sheet_name: dataframe}, column_mapping, row_counts).
    """
    # Use pandas to read all sheets at once (much faster than openpyxl iteration)
    excel_file = pd.ExcelFile(excel_path, engine='openpyxl')
    sheets_data = {}
    column_mapping = {}  # {sheet_name: {original_col: sanitized_col}}
    row_counts = {}  # {sheet_name: row_count}

    # Also use openpyxl to get structure for sheets with only formulas
    wb_structure = openpyxl.load_workbook(excel_path, read_only=True, data_only=False)

    for sheet_name in excel_file.sheet_names:
        # Read entire sheet at once
        df = pd.read_excel(excel_file, sheet_name=sheet_name, header=0, engine='openpyxl')

        # Check if the sheet is empty (all formula cells without cached values)
        if len(df) == 0:
            # Get the structure from openpyxl
            ws = wb_structure[sheet_name]
            row_count = ws.max_row - 1  # Exclude header
            col_count = ws.max_column

            # Get column headers
            headers = []
            for col_idx in range(1, col_count + 1):
                header_val = ws.cell(1, col_idx).value
                headers.append(str(header_val) if header_val else f'Column{col_idx}')

            # Store original columns for mapping
            original_cols = headers

            # Create empty DataFrame with sanitized column names
            sanitized_cols = [sanitize_column_name(h) for h in headers]
            df = pd.DataFrame(columns=sanitized_cols, index=range(row_count))
        else:
            # Store original columns and create sanitized versions
            original_cols = list(df.columns)
            sanitized_cols = [sanitize_column_name(c) for c in original_cols]

            # Rename columns to sanitized versions
            df.columns = sanitized_cols

        # Store mapping
        table_name = sheet_name.lower().replace(' ', '_')
        column_mapping[table_name] = {
            orig: sanit for orig, sanit in zip(original_cols, sanitized_cols)
        }

        df['_row'] = range(1, len(df) + 1)  # Add row index for JOINs

        # Store row count
        row_counts[table_name] = len(df)

        sheets_data[table_name] = df

        # Register with DuckDB
        conn.register(table_name, df)

    wb_structure.close()
    return sheets_data, column_mapping, row_counts


# ─────────────────────────────────────────────────────────────────
# 3. Formula Sampling & Pattern Detection
# ─────────────────────────────────────────────────────────────────

def sample_formulas(excel_path: Path, column_mapping: Dict[str, Dict[str, str]], sample_rows: int = 10) -> Dict[str, Dict[str, Tuple[str, Dict]]]:
    """
    Sample formulas from Excel to detect patterns.
    Returns {sheet_name: {sanitized_col_name: (pattern_type, params)}}
    """
    formulas_info = {}
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=False)

    # First, build column mapping for all sheets
    all_column_mappings = {}  # {sheet_name: {col_letter: sanitized_name}}

    for sheet in wb.worksheets:
        sheet_name = sheet.title.lower().replace(' ', '_')
        col_letter_to_sanitized = {}
        for col_idx in range(1, sheet.max_column + 1):
            col_letter = chr(64 + col_idx)  # A, B, C...
            header_cell = sheet.cell(1, col_idx)
            original_header = str(header_cell.value) if header_cell.value else f'Column{col_idx}'
            sanitized = column_mapping.get(sheet_name, {}).get(original_header, sanitize_column_name(original_header))
            col_letter_to_sanitized[col_letter] = sanitized
        all_column_mappings[sheet_name] = col_letter_to_sanitized

    # Now sample formulas
    for sheet in wb.worksheets:
        sheet_name = sheet.title.lower().replace(' ', '_')
        formulas_info[sheet_name] = {}

        col_letter_to_sanitized = all_column_mappings[sheet_name]

        # Sample first few data rows
        # We iterate by column first, so we can find formulas in each column
        for col_idx in range(1, sheet.max_column + 1):
            col_letter = chr(64 + col_idx)
            header_cell = sheet.cell(1, col_idx)
            original_header = str(header_cell.value) if header_cell.value else f'Column{col_idx}'
            target_col = column_mapping.get(sheet_name, {}).get(original_header, sanitize_column_name(original_header))

            # Check first few rows in this column for formulas
            for row_idx in range(2, min(2 + sample_rows, sheet.max_row + 1)):
                cell = sheet.cell(row_idx, col_idx)
                if cell.data_type == 'f':  # Formula cell
                    # Detect pattern
                    pattern = FormulaPatternDetector.detect_pattern(cell.value)
                    if pattern:
                        pattern_type, params = pattern
                        # Map column letters to sanitized names
                        if 'col1' in params:
                            params['col1_sanitized'] = col_letter_to_sanitized.get(params['col1'].upper(), params['col1'])
                        if 'col2' in params:
                            params['col2_sanitized'] = col_letter_to_sanitized.get(params['col2'].upper(), params['col2'])
                        if 'col' in params:
                            params['col_sanitized'] = col_letter_to_sanitized.get(params['col'].upper(), params['col'])
                        if 'source_col' in params:
                            # For cross-sheet references, map to the actual column in the source sheet
                            source_sheet = params['source_sheet']
                            source_col_letter = params['source_col'].upper()
                            source_mapping = all_column_mappings.get(source_sheet, {})
                            params['source_col_sanitized'] = source_mapping.get(source_col_letter, source_col_letter.lower())

                        formulas_info[sheet_name][target_col] = (pattern_type, params)
                    break  # Found formula in this column, move to next column

        # Clean up: if no formulas in a column, remove the entry
        if not formulas_info[sheet_name]:
            del formulas_info[sheet_name]

    wb.close()
    return formulas_info


# ─────────────────────────────────────────────────────────────────
# 4. Bulk Formula Evaluation with SQL
# ─────────────────────────────────────────────────────────────────

def evaluate_formulas_bulk(
    conn: duckdb.DuckDBPyConnection,
    sheets_data: Dict[str, pd.DataFrame],
    formulas_info: Dict[str, Dict[str, Tuple[str, Dict]]]
) -> Dict[str, pd.DataFrame]:
    """
    Evaluate all formulas using bulk SQL queries.
    Returns dict of updated dataframes.
    """
    result_sheets = {}

    for sheet_name, df in sheets_data.items():
        result_df = df.copy()
        sheet_formulas = formulas_info.get(sheet_name, {})

        # Process formulas one at a time, updating the DuckDB table after each
        # This allows formulas to reference computed columns from previous formulas
        for target_col, (pattern_type, params) in sheet_formulas.items():
            # Get existing columns (before adding computed one)
            # Always keep _row for JOINs
            existing_cols = ['_row'] + [c for c in result_df.columns if c != target_col and c != '_row']

            # Generate SQL for this formula pattern
            sql = FormulaPatternDetector.generate_sql(
                pattern_type, params, sheet_name, target_col, existing_cols
            )

            if sql:
                # Execute bulk SQL query
                query_result = conn.execute(sql).fetchdf()

                # Update all columns from the query result
                for col in query_result.columns:
                    if col in result_df.columns:
                        result_df[col] = query_result[col].values

                # Re-register with DuckDB so subsequent formulas can use computed values
                conn.register(sheet_name, result_df)

        result_sheets[sheet_name] = result_df

    return result_sheets


# ─────────────────────────────────────────────────────────────────
# 5. Bulk Export to Excel
# ─────────────────────────────────────────────────────────────────

def export_to_excel(
    sheets_data: Dict[str, pd.DataFrame],
    output_path: Path,
    original_path: Path
):
    """
    Export results to Excel using pandas (bulk write).
    Preserves original sheet structure.
    """
    # Read original to get column order
    original_wb = openpyxl.load_workbook(original_path, read_only=True)
    sheet_order = [s.title for s in original_wb.worksheets]
    original_wb.close()

    # Create writer with openpyxl engine
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for sheet_name in sheet_order:
            normalized_name = sheet_name.lower().replace(' ', '_')
            if normalized_name in sheets_data:
                df = sheets_data[normalized_name]

                # Remove _row column (temporary)
                df_export = df.drop(columns=['_row']) if '_row' in df.columns else df

                # Write entire sheet at once
                df_export.to_excel(writer, sheet_name=sheet_name, index=False)


# ─────────────────────────────────────────────────────────────────
# 6. Memory & Timing Measurement
# ─────────────────────────────────────────────────────────────────

def measure_memory(process: psutil.Process, stop_event: threading.Event, peak_memory: list):
    """Monitor process memory usage in background thread."""
    while not stop_event.is_set():
        try:
            mem = process.memory_info().rss / 1024 / 1024
            for child in process.children(recursive=True):
                try:
                    mem += child.memory_info().rss / 1024 / 1024
                except psutil.NoSuchProcess:
                    pass
            peak_memory[0] = max(peak_memory[0], mem)
            time.sleep(0.05)
        except psutil.NoSuchProcess:
            break


# ─────────────────────────────────────────────────────────────────
# 7. Main
# ─────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("Usage: python3 measure_duckdb_optimized.py <input_file> [output_file]")
        sys.exit(1)

    input_path = Path(sys.argv[1])
    if not input_path.exists():
        print(f"Error: File not found: {input_path}")
        sys.exit(1)

    output_path = Path(sys.argv[2]) if len(sys.argv) >= 3 else Path(f"output_{input_path.stem}_optimized.xlsx")

    process = psutil.Process()
    baseline_mem = process.memory_info().rss / 1024 / 1024

    # Start memory monitoring
    peak_memory = [baseline_mem]
    stop_event = threading.Event()
    monitor_thread = threading.Thread(target=measure_memory, args=(process, stop_event, peak_memory))
    monitor_thread.start()

    start_time = time.time()

    # Step 1: Load Excel into DuckDB (bulk read via pandas)
    conn = duckdb.connect(':memory:')
    sheets_data, column_mapping, row_counts = load_excel_to_duckdb(input_path, conn)

    # Step 2: Sample formulas to detect patterns (minimal overhead)
    formulas_info = sample_formulas(input_path, column_mapping, sample_rows=10)

    # Step 3: Evaluate formulas using bulk SQL queries
    result_sheets = evaluate_formulas_bulk(conn, sheets_data, formulas_info)

    # Step 4: Export results (bulk write via pandas)
    export_to_excel(result_sheets, output_path, input_path)

    end_time = time.time()

    # Stop monitoring
    stop_event.set()
    monitor_thread.join()

    # Calculate metrics
    total_rows = sum(len(df) for df in sheets_data.values())
    elapsed = end_time - start_time
    peak_total = peak_memory[0]
    used_mem = peak_total - baseline_mem

    output = {
        "rows": total_rows,
        "peakTotalMB": round(peak_total, 2),
        "usedMB": round(used_mem, 2),
        "baselineMB": round(baseline_mem, 2),
        "timeSeconds": round(elapsed, 3),
        "implementation": "optimized_sql"
    }

    print(json.dumps(output))
    conn.close()


if __name__ == "__main__":
    main()
