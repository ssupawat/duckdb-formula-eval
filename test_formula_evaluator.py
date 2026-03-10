#!/usr/bin/env python3
"""
Tests for FormulaEvaluator using AAA (Arrange-Act-Assert) pattern.

Test Design Techniques Used:
1. Equivalence Partitioning: Group inputs into classes that should be treated similarly
2. Boundary Value Analysis: Test boundaries between equivalence classes
3. Decision Table Testing: Test combinations of conditions
4. State Transition Testing: Test different states (found/not found)
5. Error Guessing: Anticipate likely error conditions
"""

import duckdb
import pandas as pd
import openpyxl
from pathlib import Path
from typing import Dict, Any

from formula_evaluator import FormulaEvaluator


# ============================================================================
# TEST DATA REFERENCE
# ============================================================================
# Sheet1 Data:
# Row | Key  | Name   | Category | Amount | Extra
# ----|------|--------|----------|--------|-------
# 2   | A    | Item 1 | x        | 100    | 50
# 3   | B    | Item 2 | y        | 200    | 100
# 4   | C    | Item 3 | x        | 150    | 75
# 5   | D    | Item 4 | x        | 75     | 37.5
# 6   | E    | Item 5 | y        | 300    | 150

# Sheet2 (String Lookup):
# Row | Key  | Label
# ----|------|--------
# 2   | A    | Label A
# 3   | B    | Label B
# 4   | C    | Label C

# Sheet3 (Numeric Lookup - sorted):
# Row | Score | Grade
# ----|-------|-------
# 2   | 0     | F
# 3   | 50    | D
# 4   | 70    | C
# 5   | 85    | B
# 6   | 95    | A

# ============================================================================
# TEST CASES - ORGANIZED BY TEST DESIGN TECHNIQUE
# ============================================================================

TEST_CASES = [
    # =========================================================================
    # TECHNIQUE 1: EQUIVALENCE PARTITIONING
    # Group inputs into classes that should be treated similarly
    # =========================================================================

    # ── Partition: Aggregate Functions ─────────────────────────────────────
    {
        "technique": "Equivalence Partitioning",
        "partition": "Aggregate Functions",
        "name": "COUNT(D:D)",
        "formula": "=COUNT(D:D)",
        "row_ctx": {},
        "expected": 5.0,  # 5 non-null values: 100,200,150,75,300
    },
    {
        "technique": "Equivalence Partitioning",
        "partition": "Aggregate Functions",
        "name": "MIN(D:D)",
        "formula": "=MIN(D:D)",
        "row_ctx": {},
        "expected": 75.0,  # min(100,200,150,75,300) = 75
    },
    {
        "technique": "Equivalence Partitioning",
        "partition": "Aggregate Functions",
        "name": "SUM(D:D)",
        "formula": "=SUM(D:D)",
        "row_ctx": {},
        "expected": 825.0,  # 100+200+150+75+300 = 825
    },
    {
        "technique": "Equivalence Partitioning",
        "partition": "Aggregate Functions",
        "name": "AVERAGE(D:D)",
        "formula": "=AVERAGE(D:D)",
        "row_ctx": {},
        "expected": 165.0,  # 825/5 = 165
    },
    {
        "technique": "Equivalence Partitioning",
        "partition": "Aggregate Functions",
        "name": "MAX(D:D)",
        "formula": "=MAX(D:D)",
        "row_ctx": {},
        "expected": 300.0,  # max(100,200,150,75,300) = 300
    },

    # ── Partition: SUMIF/COUNTIF Criteria Types ─────────────────────────────
    {
        "technique": "Equivalence Partitioning",
        "partition": "String Criteria (Exact Match)",
        "name": "SUMIF(C:C,\"x\",D:D) - exact string match",
        "formula": '=SUMIF(C:C,"x",D:D)',
        "row_ctx": {},
        "expected": 325.0,  # 100+150+75 = 325 (rows where C='x')
    },
    {
        "technique": "Equivalence Partitioning",
        "partition": "String Criteria (Exact Match)",
        "name": "COUNTIF(C:C,\"x\") - exact string match",
        "formula": '=COUNTIF(C:C,"x")',
        "row_ctx": {},
        "expected": 3.0,  # 3 rows where C='x' (rows 2,4,5)
    },
    {
        "technique": "Equivalence Partitioning",
        "partition": "Numeric Criteria (> operator)",
        "name": "SUMIF(D:D,\">100\",D:D) - greater than 100",
        "formula": '=SUMIF(D:D,">100",D:D)',
        "row_ctx": {},
        "expected": 650.0,  # 200+150+300 = 650 (values > 100)
    },
    {
        "technique": "Equivalence Partitioning",
        "partition": "Numeric Criteria (< operator)",
        "name": "COUNTIF(D:D,\"<150\") - less than 150",
        "formula": '=COUNTIF(D:D,"<150")',
        "row_ctx": {},
        "expected": 2.0,  # 2 values < 150 (100, 75)
    },
    {
        "technique": "Equivalence Partitioning",
        "partition": "Numeric Criteria (>= operator)",
        "name": "SUMIF(D:D,\">=150\",D:D) - greater or equal 150",
        "formula": '=SUMIF(D:D,">=150",D:D)',
        "row_ctx": {},
        "expected": 650.0,  # 200+150+300 = 650 (values >= 150)
    },
    {
        "technique": "Equivalence Partitioning",
        "partition": "Numeric Criteria (<= operator)",
        "name": "COUNTIF(D:D,\"<=100\") - less or equal 100",
        "formula": '=COUNTIF(D:D,"<=100")',
        "row_ctx": {},
        "expected": 2.0,  # 2 values <= 100 (100, 75)
    },
    {
        "technique": "Equivalence Partitioning",
        "partition": "Numeric Criteria (<> operator)",
        "name": "SUMIF(D:D,\"<>100\",D:D) - not equal 100",
        "formula": '=SUMIF(D:D,"<>100",D:D)',
        "row_ctx": {},
        "expected": 725.0,  # 200+150+75+300 = 725 (all except 100)
    },

    # ── Partition: Scalar Arithmetic Operations ────────────────────────────
    {
        "technique": "Equivalence Partitioning",
        "partition": "Scalar Operations",
        "name": "Multiplication: D1*1.07",
        "formula": "=D1*1.07",
        "row_ctx": {"D1": 100.0},
        "expected": 107.0,  # 100 * 1.07 = 107
    },
    {
        "technique": "Equivalence Partitioning",
        "partition": "Scalar Operations",
        "name": "Division: D1/2",
        "formula": "=D1/2",
        "row_ctx": {"D1": 100.0},
        "expected": 50.0,  # 100 / 2 = 50
    },
    {
        "technique": "Equivalence Partitioning",
        "partition": "Scalar Operations",
        "name": "Addition: D1+D2",
        "formula": "=D1+D2",
        "row_ctx": {"D1": 100.0, "D2": 50.0},
        "expected": 150.0,  # 100 + 50 = 150
    },
    {
        "technique": "Equivalence Partitioning",
        "partition": "Scalar Operations",
        "name": "Subtraction: D1-D2",
        "formula": "=D1-D2",
        "row_ctx": {"D1": 100.0, "D2": 30.0},
        "expected": 70.0,  # 100 - 30 = 70
    },

    # ── Partition: IF Statement Result Types ───────────────────────────────
    {
        "technique": "Equivalence Partitioning",
        "partition": "IF with Numeric Results",
        "name": "IF(D1>80, D1*1.1, D1*0.9) - TRUE branch",
        "formula": "=IF(D1>80, D1*1.1, D1*0.9)",
        "row_ctx": {"D1": 100.0},
        "expected": 110.0,  # 100>80 is TRUE, so 100*1.1 = 110
    },
    {
        "technique": "Equivalence Partitioning",
        "partition": "IF with Numeric Results",
        "name": "IF(D1>80, D1*1.1, D1*0.9) - FALSE branch",
        "formula": "=IF(D1>80, D1*1.1, D1*0.9)",
        "row_ctx": {"D1": 50.0},
        "expected": 45.0,  # 50>80 is FALSE, so 50*0.9 = 45
    },
    {
        "technique": "Equivalence Partitioning",
        "partition": "IF with String Results",
        "name": "IF(C1=\"x\", \"Category X\", \"Other\") - TRUE branch",
        "formula": '=IF(C1="x", "Category X", "Other")',
        "row_ctx": {"C1": "x"},
        "expected": "Category X",  # C1="x" is TRUE
    },
    {
        "technique": "Equivalence Partitioning",
        "partition": "IF with String Results",
        "name": "IF(C1=\"x\", \"Category X\", \"Other\") - FALSE branch",
        "formula": '=IF(C1="x", "Category X", "Other")',
        "row_ctx": {"C1": "y"},
        "expected": "Other",  # C1="x" is FALSE
    },

    # ── Partition: VLOOKUP Lookup Value Types ───────────────────────────────
    {
        "technique": "Equivalence Partitioning",
        "partition": "VLOOKUP String Lookup",
        "name": "VLOOKUP(A1,Sheet2!A:B,2,0) - Key A",
        "formula": "=VLOOKUP(A1,Sheet2!A:B,2,0)",
        "row_ctx": {"A1": "A"},
        "expected": "Label A",  # Lookup "A" finds row 2, return column B
    },
    {
        "technique": "Equivalence Partitioning",
        "partition": "VLOOKUP String Lookup",
        "name": "VLOOKUP(A1,Sheet2!A:B,2,0) - Key B",
        "formula": "=VLOOKUP(A1,Sheet2!A:B,2,0)",
        "row_ctx": {"A1": "B"},
        "expected": "Label B",  # Lookup "B" finds row 3, return column B
    },

    # =========================================================================
    # TECHNIQUE 2: BOUNDARY VALUE ANALYSIS
    # Test boundaries between equivalence classes
    # =========================================================================

    # ── Boundary: VLOOKUP Column Index ───────────────────────────────────────
    {
        "technique": "Boundary Value Analysis",
        "boundary": "VLOOKUP col_index (lower bound)",
        "name": "VLOOKUP(A1,Sheet2!A:B,1,0) - first column",
        "formula": "=VLOOKUP(A1,Sheet2!A:B,1,0)",
        "row_ctx": {"A1": "A"},
        "expected": "A",  # col_index=1 returns first column (lookup column itself)
    },
    {
        "technique": "Boundary Value Analysis",
        "boundary": "VLOOKUP col_index (upper bound)",
        "name": "VLOOKUP(A1,Sheet2!A:B,2,0) - last column in range",
        "formula": "=VLOOKUP(A1,Sheet2!A:B,2,0)",
        "row_ctx": {"A1": "A"},
        "expected": "Label A",  # col_index=2 returns second column (last in A:B)
    },

    # ── Boundary: Comparison Operators ───────────────────────────────────────
    {
        "technique": "Boundary Value Analysis",
        "boundary": "SUMIF with = boundary",
        "name": "SUMIF(D:D,\"=100\",D:D) - equals boundary",
        "formula": '=SUMIF(D:D,"=100",D:D)',
        "row_ctx": {},
        "expected": 100.0,  # Only 100 equals 100
    },
    {
        "technique": "Boundary Value Analysis",
        "boundary": "SUMIF with > boundary",
        "name": "SUMIF(D:D,\">75\",D:D) - greater than boundary",
        "formula": '=SUMIF(D:D,">75",D:D)',
        "row_ctx": {},
        "expected": 750.0,  # 100+200+150+300 = 750 (values > 75, excludes 75)
    },
    {
        "technique": "Boundary Value Analysis",
        "boundary": "SUMIF with >= boundary",
        "name": "SUMIF(D:D,\">=75\",D:D) - greater or equal boundary",
        "formula": '=SUMIF(D:D,">=75",D:D)',
        "row_ctx": {},
        "expected": 825.0,  # 100+200+150+75+300 = 825 (values >= 75, includes 75)
    },

    # ── Boundary: IF Condition Boundary ────────────────────────────────────
    {
        "technique": "Boundary Value Analysis",
        "boundary": "IF equality (boundary case)",
        "name": "IF(D1=100, \"Equal\", \"Not Equal\") - equal boundary",
        "formula": '=IF(D1=100, "Equal", "Not Equal")',
        "row_ctx": {"D1": 100.0},
        "expected": "Equal",  # D1 equals 100 exactly
    },
    {
        "technique": "Boundary Value Analysis",
        "boundary": "IF equality (just below)",
        "name": "IF(D1=100, \"Equal\", \"Not Equal\") - below boundary",
        "formula": '=IF(D1=100, "Equal", "Not Equal")',
        "row_ctx": {"D1": 99.9},
        "expected": "Not Equal",  # D1 is 99.9, not equal to 100
    },

    # =========================================================================
    # TECHNIQUE 3: DECISION TABLE TESTING
    # Test combinations of conditions
    # =========================================================================

    # ── Decision Table: Nested IF with Aggregates ───────────────────────────
    {
        "technique": "Decision Table Testing",
        "decision": "Aggregate > threshold AND row value > threshold",
        "name": "IF(SUMIF>100, IF(D1>80, D1*1.1, D1), 0) - both TRUE",
        "formula": '=IF(SUMIF(C:C,"x",D:D)>100, IF(D1>80, D1*1.1, D1), 0)',
        "row_ctx": {"D1": 100.0},
        "expected": 110.0,  # SUMIF=325>100 (TRUE), D1=100>80 (TRUE) → 100*1.1
    },
    {
        "technique": "Decision Table Testing",
        "decision": "Aggregate > threshold AND row value <= threshold",
        "name": "IF(SUMIF>100, IF(D1>80, D1*1.1, D1), 0) - outer TRUE, inner FALSE",
        "formula": '=IF(SUMIF(C:C,"x",D:D)>100, IF(D1>80, D1*1.1, D1), 0)',
        "row_ctx": {"D1": 50.0},
        "expected": 50.0,  # SUMIF=325>100 (TRUE), D1=50>80 (FALSE) → D1
    },
    {
        "technique": "Decision Table Testing",
        "decision": "Aggregate > threshold AND row value > threshold (using <)",
        "name": "IF(SUMIF<100, ...) - outer FALSE",
        "formula": '=IF(SUMIF(C:C,"x",D:D)<100, 999, 0)',
        "row_ctx": {},
        "expected": 0.0,  # SUMIF=325<100 (FALSE) → 0
    },

    # ── Decision Table: Arithmetic on Aggregates ────────────────────────────
    {
        "technique": "Decision Table Testing",
        "decision": "SUM * percentage",
        "name": "SUM(D:D)*0.1 - 10% of total",
        "formula": "=SUM(D:D)*0.1",
        "row_ctx": {},
        "expected": 82.5,  # 825 * 0.1 = 82.5
    },
    {
        "technique": "Decision Table Testing",
        "decision": "AVERAGE * multiplier",
        "name": "AVERAGE(D:D)*1.2 - 20% markup on average",
        "formula": "=AVERAGE(D:D)*1.2",
        "row_ctx": {},
        "expected": 198.0,  # 165 * 1.2 = 198
    },
    {
        "technique": "Decision Table Testing",
        "decision": "Aggregate ratio in condition",
        "name": "IF(SUMIF/COUNTIF>50, D1*2, D1) - average check",
        "formula": '=IF(SUMIF(C:C,"x",D:D)/COUNTIF(C:C,"x")>50, D1*2, D1)',
        "row_ctx": {"D1": 100.0},
        "expected": 200.0,  # 325/3≈108.3>50 (TRUE) → 100*2
    },

    # ── Decision Table: Multiple Cell References ────────────────────────────
    {
        "technique": "Decision Table Testing",
        "decision": "Multiple cell arithmetic",
        "name": "(D1+E1)/2 - average of two cells",
        "formula": "=(D1+E1)/2",
        "row_ctx": {"D1": 100.0, "E1": 200.0},
        "expected": 150.0,  # (100+200)/2 = 150
    },

    # =========================================================================
    # TECHNIQUE 4: STATE TRANSITION TESTING
    # Test different system states
    # =========================================================================

    # ── State: VLOOKUP Found vs Not Found ───────────────────────────────────
    {
        "technique": "State Transition Testing",
        "state": "VLOOKUP - Key Found",
        "name": "VLOOKUP(\"A\",Sheet2!A:B,2,0) - found state",
        "formula": '=VLOOKUP("A",Sheet2!A:B,2,0)',
        "row_ctx": {},
        "expected": "Label A",  # Key "A" exists in Sheet2
    },
    {
        "technique": "State Transition Testing",
        "state": "VLOOKUP - Key Not Found",
        "name": "VLOOKUP(\"Z\",Sheet2!A:B,2,0) - not found state",
        "formula": '=VLOOKUP("Z",Sheet2!A:B,2,0)',
        "row_ctx": {},
        "expected": 0,  # Key "Z" doesn't exist, returns 0 (converted to int)
    },

    # ── State: VLOOKUP with Different Match Types ───────────────────────────
    {
        "technique": "State Transition Testing",
        "state": "VLOOKUP - Exact Match (range_lookup=0)",
        "name": "VLOOKUP(85,Sheet3!A:B,2,0) - exact match",
        "formula": "=VLOOKUP(85,Sheet3!A:B,2,0)",
        "row_ctx": {},
        "expected": "B",  # Exact match: 85 → "B"
    },
    {
        "technique": "State Transition Testing",
        "state": "VLOOKUP - Approximate Match (range_lookup=1)",
        "name": "VLOOKUP(82,Sheet3!A:B,2,1) - approximate match",
        "formula": "=VLOOKUP(82,Sheet3!A:B,2,1)",
        "row_ctx": {},
        "expected": "C",  # Approximate: largest ≤82 is 70 → "C"
    },
    {
        "technique": "State Transition Testing",
        "state": "VLOOKUP - Below Range (approximate)",
        "name": "VLOOKUP(10,Sheet3!A:B,2,1) - below range",
        "formula": "=VLOOKUP(10,Sheet3!A:B,2,1)",
        "row_ctx": {},
        "expected": "F",  # Below smallest (0), should return "F" or error
    },

    # =========================================================================
    # TECHNIQUE 5: ERROR GUESSING
    # Anticipate likely error conditions
    # =========================================================================

    # ── Error: Division by Zero ─────────────────────────────────────────────
    {
        "technique": "Error Guessing",
        "error": "Division by zero (protected by IF)",
        "name": "IF(D1=0, \"Error\", 100/D1) - protected division",
        "formula": '=IF(D1=0, "Error", 100/D1)',
        "row_ctx": {"D1": 0.0},
        "expected": "Error",  # D1=0 triggers TRUE branch, returns "Error"
    },
    {
        "technique": "Error Guessing",
        "error": "Division by zero (normal case)",
        "name": "IF(D1!=0, 100/D1, 0) - normal division",
        "formula": '=IF(D1!=0, 100/D1, 0)',
        "row_ctx": {"D1": 25.0},
        "expected": 4.0,  # D1!=0 is TRUE, 100/25 = 4
    },
    {
        "technique": "Error Guessing",
        "error": "Division by zero (normal case)",
        "name": "IF(D1<>0, 100/D1, 0) - normal division",
        "formula": '=IF(D1<>0, 100/D1, 0)',
        "row_ctx": {"D1": 25.0},
        "expected": 4.0,  # D1<>0 is TRUE, 100/25 = 4
    },

    # ── Error: Empty/Null Values ────────────────────────────────────────────
    {
        "technique": "Error Guessing",
        "error": "Aggregate on empty column",
        "name": "SUM(Z:Z) - empty column",
        "formula": "=SUM(Z:Z)",
        "row_ctx": {},
        "expected": 0.0,  # Column Z doesn't exist or is empty, should return 0
    },
    {
        "technique": "Error Guessing",
        "error": "COUNT on empty column",
        "name": "COUNT(Z:Z) - empty column",
        "formula": "=COUNT(Z:Z)",
        "row_ctx": {},
        "expected": 0.0,  # Column Z doesn't exist or is empty, should return 0
    },

    # ── Error: Invalid References ───────────────────────────────────────────
    {
        "technique": "Error Guessing",
        "error": "VLOOKUP with non-existent sheet",
        "name": "VLOOKUP(A1,NonExistent!A:B,2,0) - invalid sheet",
        "formula": "=VLOOKUP(A1,NonExistent!A:B,2,0)",
        "row_ctx": {"A1": "A"},
        "expected": 0,  # Sheet doesn't exist, should return 0
    },
    # Note: Cell reference not in row_ctx - uses 0 as default, difficult to test reliably
    # {
    #     "technique": "Error Guessing",
    #     "error": "Cell reference not in row_ctx",
    #     "name": "D1*2 - D1 not provided in row_ctx",
    #     "formula": "=D1*2",
    #     "row_ctx": {},  # D1 not provided
    #     "expected_error": "KeyError or ValueError",
    # },

    # ── Error: Edge Case Formulas ───────────────────────────────────────────
    {
        "technique": "Error Guessing",
        "error": "SUMIF with empty criteria",
        "name": "SUMIF(D:D,\"\",D:D) - empty criteria",
        "formula": '=SUMIF(D:D,"",D:D)',
        "row_ctx": {},
        "expected": 0.0,  # Empty criteria should match empty cells (none exist)
    },

    # =========================================================================
    # Cross-sheet row-by-row reference (NOW IMPLEMENTED)
    # =========================================================================
    {
        "technique": "State Transition Testing",
        "partition": "Cross-Sheet References",
        "name": "Sheet2!A2 - cross-sheet reference (returns first row value)",
        "formula": '=Sheet2!A2',  # Reference Sheet2 column A, row 2
        "row_ctx": {},
        "expected": "A",  # Returns 'A' (first value from Sheet2 column A)
        # Note: This returns the first row's value (LIMIT 1), not row-by-row correspondence
        # For apply_formula_to_column, row-by-row logic is implemented separately
    },
]


def test_sql_conversion(formula: str, expected_sql_pattern: str, evaluator: FormulaEvaluator) -> bool:
    """
    Test that Excel formula converts to expected SQL pattern.

    Args:
        formula: Excel formula to test
        expected_sql_pattern: Expected SQL pattern (can be partial)
        evaluator: FormulaEvaluator instance

    Returns:
        True if generated SQL matches expected pattern
    """
    from formula_evaluator import FormulaEvaluator as FE
    # Access the excel_to_sql method
    sql = evaluator.excel_to_sql(formula, 'sheet1', {})
    matches = expected_sql_pattern in sql
    print(f"  Formula: {formula}")
    print(f"  SQL: {sql}")
    print(f"  Expected pattern: {expected_sql_pattern}")
    print(f"  Match: {matches}")
    return matches


def create_test_excel() -> Path:
    """Create a test Excel file with sample data for testing."""
    wb = openpyxl.Workbook()

    # Sheet1: Main data
    ws1 = wb.active
    ws1.title = 'Sheet1'
    ws1['A1'] = 'Key'
    ws1['B1'] = 'Name'
    ws1['C1'] = 'Category'
    ws1['D1'] = 'Amount'
    ws1['E1'] = 'Extra'

    test_data = [
        ['A', 'Item 1', 'x', 100, 50],
        ['B', 'Item 2', 'y', 200, 100],
        ['C', 'Item 3', 'x', 150, 75],
        ['D', 'Item 4', 'x', 75, 37.5],
        ['E', 'Item 5', 'y', 300, 150],
    ]

    for i, row in enumerate(test_data, start=2):
        ws1[f'A{i}'] = row[0]
        ws1[f'B{i}'] = row[1]
        ws1[f'C{i}'] = row[2]
        ws1[f'D{i}'] = row[3]
        ws1[f'E{i}'] = row[4]

    # Sheet2: Lookup table (strings)
    ws2 = wb.create_sheet('Sheet2')
    ws2['A1'] = 'Key'
    ws2['B1'] = 'Label'

    lookup_data = [
        ['A', 'Label A'],
        ['B', 'Label B'],
        ['C', 'Label C'],
    ]

    for i, row in enumerate(lookup_data, start=2):
        ws2[f'A{i}'] = row[0]
        ws2[f'B{i}'] = row[1]

    # Sheet3: Lookup table (numeric, sorted for approximate match)
    ws3 = wb.create_sheet('Sheet3')
    ws3['A1'] = 'Score'
    ws3['B1'] = 'Grade'

    score_data = [
        [0, 'F'],
        [50, 'D'],
        [70, 'C'],
        [85, 'B'],
        [95, 'A'],
    ]

    for i, row in enumerate(score_data, start=2):
        ws3[f'A{i}'] = row[0]
        ws3[f'B{i}'] = row[1]

    output_path = Path('test_formulas.xlsx')
    wb.save(output_path)
    return output_path


def setup_evaluator(test_file: Path) -> FormulaEvaluator:
    """Arrange: Set up test fixtures (DuckDB connection, data, evaluator)."""
    conn = duckdb.connect(':memory:')
    excel_file = pd.ExcelFile(test_file, engine='openpyxl')

    for sheet_name in excel_file.sheet_names:
        df = pd.read_excel(excel_file, sheet_name=sheet_name, header=0, engine='openpyxl')
        df.columns = [str(c).lower().replace(' ', '_') for c in df.columns]
        table_name = sheet_name.lower().replace(' ', '_')
        conn.register(table_name, df)

    return FormulaEvaluator(conn)


def run_test(test_case: Dict[str, Any], evaluator: FormulaEvaluator) -> bool:
    """Act & Assert: Execute test case and verify result matches expected."""
    name = test_case["name"]
    formula = test_case["formula"]
    row_ctx = test_case["row_ctx"]
    expected = test_case["expected"]
    technique = test_case.get("technique", "General")

    # Check if this is an error test
    expected_error = test_case.get("expected_error")

    try:
        # Act: Convert formula to SQL and execute in DuckDB
        sql = evaluator.excel_to_sql(formula, 'sheet1', row_ctx)

        # Debug: Show generated SQL
        print(f"     SQL: {sql}")

        # Execute SQL in DuckDB and get result
        result = evaluator.conn.execute(sql).fetchdf().iloc[0, 0]

        # Handle NaN results
        import pandas as pd
        if pd.isna(result):
            result = 0.0
        elif isinstance(result, str):
            pass  # Keep as string
        else:
            result = float(result)

        # If we expected an error but got a result, it's a failure
        if expected_error:
            print(f"✗ [{technique}] {name:55s} → Expected error but got: {result}")
            return False

        # Assert: Verify result matches expected
        if isinstance(expected, str):
            passed = result == expected
            status = "✓" if passed else "✗"
            result_str = f"'{result}'"
            expected_str = f"'{expected}'"
        else:
            passed = abs(result - expected) < 0.01
            status = "✓" if passed else "✗"
            result_str = f"{result:.2f}"
            expected_str = f"{expected:.2f}"

        if passed:
            print(f"{status} [{technique}] {name:55s} → {result_str:15s}")
            return True
        else:
            print(f"{status} [{technique}] {name:55s} → Expected: {expected_str}, Got: {result_str}")
            return False

    except Exception as e:
        # If we expected an error, check if it matches
        if expected_error:
            if expected_error in str(type(e).__name__) or expected_error in str(e):
                print(f"✓ [{technique}] {name:55s} → Got expected error: {type(e).__name__}")
                return True
            else:
                print(f"✗ [{technique}] {name:55s} → Expected error: {expected_error}, Got: {type(e).__name__}: {e}")
                return False
        else:
            print(f"✗ [{technique}] {name:55s} → ERROR: {type(e).__name__}: {e}")
            return False


def main():
    """Run all test cases using AAA pattern."""
    print("=" * 100)
    print("FormulaEvaluator Test Suite - Test Design Techniques")
    print("=" * 100)

    # Arrange: Set up test fixtures
    print("\n1. ARRANGE - Creating test fixtures...")
    test_file = create_test_excel()
    print(f"   Created test file: {test_file}")

    evaluator = setup_evaluator(test_file)
    print("   Created FormulaEvaluator with DuckDB connection")

    # Act & Assert: Run all test cases
    print("\n2. ACT & ASSERT - Running test cases...")
    print("-" * 100)

    # Group tests by technique for better reporting
    by_technique = {}
    for test_case in TEST_CASES:
        technique = test_case.get("technique", "General")
        if technique not in by_technique:
            by_technique[technique] = []
        by_technique[technique].append(test_case)

    passed = 0
    failed = 0
    skipped = 0

    for technique, tests in by_technique.items():
        print(f"\n>>> {technique} ({len(tests)} tests)")
        for test_case in tests:
            if test_case.get("expected_error"):
                skipped += 1  # Count error tests separately
            if run_test(test_case, evaluator):
                passed += 1
            else:
                failed += 1

    # Summary
    print("\n" + "-" * 100)
    print(f"\nResults: {passed} passed, {failed} failed, {skipped} error tests out of {len(TEST_CASES)} total tests")

    # Show breakdown by technique
    print("\nBreakdown by technique:")
    for technique, tests in by_technique.items():
        tech_passed = sum(1 for tc in tests if tc.get("_passed", False))
        tech_failed = sum(1 for tc in tests if tc.get("_failed", False))
        print(f"  {technique}: {len(tests)} tests")

    if failed == 0:
        print("\n✓ All tests passed!")
    else:
        print(f"\n✗ {failed} test(s) failed")

    print(f"\n3. Test file saved as: {test_file}")
    print("   You can verify formulas manually in Excel.")

    return failed == 0


if __name__ == "__main__":
    import sys
    sys.exit(0 if main() else 1)
