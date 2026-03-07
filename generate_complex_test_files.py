#!/usr/bin/env python3
"""
Generate Excel files with complex formulas for benchmarking.
Tests formulas that require Phase 2 scalar evaluation (numexpr).
"""

import openpyxl
from pathlib import Path

def create_single_sheet_test(rows: int, output_path: Path):
    """Create single sheet with IF formulas."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    # Headers
    ws['A1'] = 'Category'
    ws['B1'] = 'Value'
    ws['C1'] = 'IF_Result'
    ws['D1'] = 'Condition'

    # Data with formulas
    for i in range(2, rows + 2):
        category = 'x' if i % 2 == 0 else 'y'
        value = i * 10
        ws[f'A{i}'] = category
        ws[f'B{i}'] = value
        ws[f'C{i}'] = f'=IF(A{i}="x", B{i}*1.1, B{i}*0.9)'
        ws[f'D{i}'] = value if category == 'x' else value // 2

    wb.save(output_path)

def create_two_sheet_test(rows: int, output_path: Path):
    """Create two-sheet file for VLOOKUP tests."""
    wb = openpyxl.Workbook()

    # Sheet1: Main data
    ws1 = wb.active
    ws1.title = 'Sheet1'
    ws1['A1'] = 'Key'
    ws1['B1'] = 'Value'
    ws1['C1'] = 'VLOOKUP_Result'

    for i in range(2, rows + 2):
        key = chr(65 + (i % 3))  # A, B, C
        ws1[f'A{i}'] = key
        ws1[f'B{i}'] = i * 10
        ws1[f'C{i}'] = f'=VLOOKUP(A{i},Sheet2!A:B,2,0)'

    # Sheet2: Lookup table
    ws2 = wb.create_sheet('Sheet2')
    ws2['A1'] = 'Key'
    ws2['B1'] = 'Label'

    lookup_data = [['A', 'Label A'], ['B', 'Label B'], ['C', 'Label C']]
    for i, row in enumerate(lookup_data, start=2):
        ws2[f'A{i}'] = row[0]
        ws2[f'B{i}'] = row[1]

    wb.save(output_path)

def main():
    sizes = [10000, 50000, 100000]
    output_dir = Path('test_files_complex')
    output_dir.mkdir(exist_ok=True)

    for n in sizes:
        create_single_sheet_test(n, output_dir / f'complex_{n}.xlsx')
        create_two_sheet_test(n, output_dir / f'complex_2sheet_{n}.xlsx')
        print(f"Generated complex_{n}.xlsx and complex_2sheet_{n}.xlsx")

if __name__ == '__main__':
    main()
