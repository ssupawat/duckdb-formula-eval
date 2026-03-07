#!/usr/bin/env python3
"""
Generate test Excel files for FormulaEvaluator.

Test formats:
- Simple: A=i, B=i*2, C=A{i}+B{i}
- Two-sheet: Sheet1 A=i, Sheet2 A=Sheet1!A{i}, B=A{i}*2
- Complex IF: IF(A{i}="x", B{i}*1.1, B{i}*0.9)
- Complex VLOOKUP: VLOOKUP(A{i},Sheet2!A:B,2,0)

Usage:
    python3 generate_test_files.py              # Generate all simple test files
    python3 generate_test_files.py 10000        # Generate simple 10K test
    python3 generate_test_files.py --complex    # Generate all complex test files
    python3 generate_test_files.py --complex 10000  # Generate complex 10K test
"""

import openpyxl
import sys
import argparse
from pathlib import Path


def create_standard_test(rows: int, output_path: Path):
    """Create standard single-sheet test (simple formulas)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    # Header
    ws['A1'] = 'A'
    ws['B1'] = 'B'
    ws['C1'] = 'C = A + B'

    # Data with formulas
    for i in range(2, rows + 2):
        ws[f'A{i}'] = i
        ws[f'B{i}'] = i * 2
        ws[f'C{i}'] = f'=A{i}+B{i}'

    wb.save(output_path)
    print(f'Generated {output_path.name}')


def create_two_sheet_test(rows: int, output_path: Path):
    """Create two-sheet test with cross-sheet references."""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = 'Sheet1'

    # Sheet1: Source data
    ws1['A1'] = 'Value'
    for i in range(2, rows + 2):
        ws1[f'A{i}'] = i

    # Sheet2: References Sheet1
    ws2 = wb.create_sheet('Sheet2')
    ws2['A1'] = 'From Sheet1'
    ws2['B1'] = 'Doubled'

    for i in range(2, rows + 2):
        ws2[f'A{i}'] = f'=Sheet1!A{i}'
        ws2[f'B{i}'] = f'=A{i}*2'

    wb.save(output_path)
    print(f'Generated {output_path.name}')


def create_complex_if_test(rows: int, output_path: Path):
    """Create single sheet with IF formulas (complex formulas)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    # Headers
    ws['A1'] = 'Category'
    ws['B1'] = 'Value'
    ws['C1'] = 'IF_Result'
    ws['D1'] = 'Condition'

    # Data with formulas
    for i in range(2, rows + 2):
        category = 'x' if i % 2 == 0 else 'y'
        value = i * 10
        ws[f'A{i}'] = category
        ws[f'B{i}'] = value
        ws[f'C{i}'] = f'=IF(A{i}="x", B{i}*1.1, B{i}*0.9)'
        ws[f'D{i}'] = value if category == 'x' else value // 2

    wb.save(output_path)
    print(f'Generated {output_path.name}')


def create_complex_vlookup_test(rows: int, output_path: Path):
    """Create two-sheet file for VLOOKUP tests (complex formulas)."""
    wb = openpyxl.Workbook()

    # Sheet1: Main data
    ws1 = wb.active
    ws1.title = 'Sheet1'
    ws1['A1'] = 'Key'
    ws1['B1'] = 'Value'
    ws1['C1'] = 'VLOOKUP_Result'

    for i in range(2, rows + 2):
        key = chr(65 + (i % 3))  # A, B, C
        ws1[f'A{i}'] = key
        ws1[f'B{i}'] = i * 10
        ws1[f'C{i}'] = f'=VLOOKUP(A{i},Sheet2!A:B,2,0)'

    # Sheet2: Lookup table
    ws2 = wb.create_sheet('Sheet2')
    ws2['A1'] = 'Key'
    ws2['B1'] = 'Label'

    lookup_data = [['A', 'Label A'], ['B', 'Label B'], ['C', 'Label C']]
    for i, row in enumerate(lookup_data, start=2):
        ws2[f'A{i}'] = row[0]
        ws2[f'B{i}'] = row[1]

    wb.save(output_path)
    print(f'Generated {output_path.name}')


def main():
    parser = argparse.ArgumentParser(
        description='Generate test Excel files for FormulaEvaluator',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
    python3 generate_test_files.py              # Generate all simple test files
    python3 generate_test_files.py 10000        # Generate simple 10K test
    python3 generate_test_files.py --complex    # Generate all complex test files
    python3 generate_test_files.py --complex 10000  # Generate complex 10K test
        """
    )
    parser.add_argument('rows', nargs='?', type=int, help='Number of rows to generate')
    parser.add_argument('--complex', action='store_true', help='Generate complex formula test files')
    args = parser.parse_args()

    # Determine output directory
    test_dir = Path('test_files')
    test_dir.mkdir(exist_ok=True)

    if args.complex:
        # Complex test files
        complex_sizes = [10000, 50000, 100000]
        rows = args.rows if args.rows else complex_sizes

        if isinstance(rows, int):
            create_complex_if_test(rows, test_dir / f'complex_{rows}.xlsx')
            create_complex_vlookup_test(rows, test_dir / f'complex_2sheet_{rows}.xlsx')
        else:
            for n in rows:
                create_complex_if_test(n, test_dir / f'complex_{n}.xlsx')
                create_complex_vlookup_test(n, test_dir / f'complex_2sheet_{n}.xlsx')
    else:
        # Simple test files
        simple_sizes = [10000, 50000, 100000, 200000]
        two_sheet_sizes = [10000, 100000, 500000]
        rows = args.rows if args.rows else simple_sizes

        if isinstance(rows, int):
            create_standard_test(rows, test_dir / f'simple_{rows}.xlsx')
            create_two_sheet_test(rows, test_dir / f'simple_2sheet_{rows}.xlsx')
        else:
            for n in rows:
                create_standard_test(n, test_dir / f'simple_{n}.xlsx')

            for n in two_sheet_sizes:
                create_two_sheet_test(n, test_dir / f'simple_2sheet_{n}.xlsx')


if __name__ == '__main__':
    main()
