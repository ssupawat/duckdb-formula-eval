#!/usr/bin/env python3
"""
Benchmark complex formulas using FormulaEvaluator.
Measures performance of numexpr scalar evaluation.
"""

import sys
import time
import psutil
import duckdb
import pandas as pd
import json
import openpyxl
from pathlib import Path

# Add FormulaEvaluator from test_formula_evaluator.py
sys.path.insert(0, str(Path(__file__).parent))
from test_formula_evaluator import FormulaEvaluator

def get_memory_mb():
    return psutil.Process().memory_info().rss / 1024 / 1024

def adjust_formula_row_numbers(formula: str, from_row: int, to_row: int) -> str:
    """Adjust row numbers in an Excel formula from one row to another."""
    import re
    # Match cell references like A2, B2, AA2, etc. (column letters + row number)
    # Use word boundary to avoid partial matches
    def replace_ref(match):
        col = match.group(1)
        row = int(match.group(2))
        if row == from_row:
            return f"{col}{to_row}"
        return match.group(0)

    # Pattern: column letters (A-Z, one or more) followed by row number
    return re.sub(r'\b([A-Z]+)(\d+)\b', replace_ref, formula)

def benchmark_complex_formulas(excel_path: str):
    """Benchmark complex formula evaluation."""
    conn = duckdb.connect(':memory:')

    # Load Excel with openpyxl to get formulas
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    sheets_data = {}
    formula_info = {}  # Store formula patterns per sheet

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        table_name = sheet_name.lower().replace(' ', '_')

        # Get all data first (as pandas df for duckdb)
        data_rows = []
        for row in ws.iter_rows(values_only=True):
            data_rows.append(row)

        df = pd.DataFrame(data_rows[1:], columns=data_rows[0]) if data_rows else pd.DataFrame()
        df.columns = [str(c).lower().replace(' ', '_') for c in df.columns]
        sheets_data[table_name] = df
        conn.register(table_name, df)

        # Find formula column and template from row 2 (first data row)
        formula_col = None
        formula_template = None
        for col_idx, cell in enumerate(ws[2], start=0):
            if cell.data_type == 'f':  # formula type
                formula_col = col_idx
                formula_template = cell.value
                break

        if formula_col is not None:
            formula_info[table_name] = {
                'col_idx': formula_col,
                'template': formula_template
            }

    evaluator = FormulaEvaluator(conn, sheets_data)

    # Get the sheet with formulas
    sheet_name = list(formula_info.keys())[0]
    formula_col_idx = formula_info[sheet_name]['col_idx']
    formula_template = formula_info[sheet_name]['template']
    df = sheets_data[sheet_name]

    # Get column letter for formula column
    formula_col_letter = chr(65 + formula_col_idx)  # A, B, C, etc.

    # Benchmark
    mem_start = get_memory_mb()
    time_start = time.perf_counter()

    results = []
    for idx in range(len(df)):
        # Adjust formula row numbers for current row
        # Template is from row 2, current row is idx + 2
        current_formula = adjust_formula_row_numbers(formula_template, 2, idx + 2)

        row_ctx = {}
        # Build row context from all columns
        for col_idx, col in enumerate(df.columns):
            if col_idx != formula_col_idx:
                cell_ref = f"{chr(65 + col_idx)}{idx + 2}"
                row_ctx[cell_ref] = df[col].iloc[idx]

        result = evaluator.evaluate_formula(current_formula, sheet_name, row_ctx)
        results.append(result)

    time_end = time.perf_counter()
    mem_end = get_memory_mb()
    mem_peak = mem_end - mem_start

    return {
        'rows': len(df),
        'timeSeconds': time_end - time_start,
        'peakMemoryMb': mem_peak
    }

def main():
    if len(sys.argv) < 2:
        print("Usage: python3 measure_complex_formulas.py <excel_file>")
        sys.exit(1)

    result = benchmark_complex_formulas(sys.argv[1])
    print(json.dumps(result))

if __name__ == '__main__':
    main()
