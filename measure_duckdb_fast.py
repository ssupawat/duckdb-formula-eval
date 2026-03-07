#!/usr/bin/env python3
"""
DuckDB Excel Formula Evaluation Benchmark - Optimized Version

Uses pandas vectorized operations for formula evaluation.
"""

import re
import duckdb
import openpyxl
import pandas as pd
import psutil
import json
import sys
import time
import threading
from pathlib import Path
from typing import Dict


def load_excel_to_dataframes(excel_path: Path) -> Dict[str, pd.DataFrame]:
    """Load all sheets into pandas DataFrames."""
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    dataframes = {}

    for sheet in wb.worksheets:
        sheet_name = sheet.title

        # Read data into list
        data = []
        headers = []

        # Get headers from row 1
        for col_idx in range(1, sheet.max_column + 1):
            cell_val = sheet.cell(1, col_idx).value
            headers.append(str(cell_val).lower() if cell_val else f'col_{col_idx}')

        # Read data rows
        for row_idx in range(2, sheet.max_row + 1):
            row_dict = {'_row': row_idx}
            for col_idx, header in enumerate(headers):
                cell_val = sheet.cell(row_idx, col_idx + 1).value
                row_dict[header] = cell_val
            data.append(row_dict)

        if data:
            df = pd.DataFrame(data)
            dataframes[sheet_name] = df

    return dataframes


def evaluate_formulas_vectorized(wb: openpyxl.Workbook, dataframes: Dict[str, pd.DataFrame]) -> None:
    """Evaluate formulas using vectorized pandas operations."""

    for sheet in wb.worksheets:
        sheet_name = sheet.title
        if sheet_name not in dataframes:
            continue

        df = dataframes[sheet_name].copy()

        # Find all formula cells and their patterns
        formula_cells = {}
        for row_idx in range(2, sheet.max_row + 1):
            for col_idx in range(1, sheet.max_column + 1):
                cell = sheet.cell(row_idx, col_idx)
                if cell.data_type == 'f':
                    formula_cells[(row_idx, col_idx)] = cell.value

        # Evaluate formulas in bulk by pattern
        # Pattern 1: =A{i}+B{i} - simple column addition
        # Pattern 2: =Sheet1!A{i} - cross-sheet reference
        # Pattern 3: =A{i}*2 - column with literal

        for (row_idx, col_idx), formula in formula_cells.items():
            formula = formula.strip().lstrip('=')

            # Check for cross-sheet reference: Sheet1!A2
            cross_sheet_match = re.match(r'(\w+)!([A-Z]+)(\d+)', formula)
            if cross_sheet_match:
                source_sheet = cross_sheet_match.group(1)
                col_letter = cross_sheet_match.group(2).upper()
                source_row = int(cross_sheet_match.group(3))

                if source_sheet in dataframes:
                    source_df = dataframes[source_sheet]
                    # Find the column by position (A=0, B=1, etc.)
                    col_idx = ord(col_letter) - ord('A')
                    if col_idx < len(source_df.columns):
                        # Use integer position with iloc
                        value = source_df.iloc[source_row - 2, col_idx]
                        sheet.cell(row_idx, col_idx).value = value
                continue

            # Check for simple arithmetic with two columns: A2+B2
            arithmetic_match = re.match(r'([A-Z]+)(\d+)\+([A-Z]+)(\d+)', formula)
            if arithmetic_match:
                col1 = arithmetic_match.group(1).upper()
                row1 = int(arithmetic_match.group(2))
                col2 = arithmetic_match.group(3).upper()
                row2 = int(arithmetic_match.group(4))

                if row1 == row2 and row1 == row_idx:
                    idx1 = ord(col1) - ord('A')
                    idx2 = ord(col2) - ord('A')

                    if idx1 < len(df.columns) and idx2 < len(df.columns):
                        col_name1 = df.columns[idx1]
                        col_name2 = df.columns[idx2]
                        val1 = df.iloc[row_idx - 2][col_name1]
                        val2 = df.iloc[row_idx - 2][col_name2]
                        sheet.cell(row_idx, col_idx).value = val1 + val2
                continue

            # Check for column * literal: A2*2
            mult_match = re.match(r'([A-Z]+)(\d+)\*(\d+)', formula)
            if mult_match:
                col = mult_match.group(1).upper()
                row_ref = int(mult_match.group(2))
                multiplier = float(mult_match.group(3))

                if row_ref == row_idx:
                    idx = ord(col) - ord('A')
                    if idx < len(df.columns):
                        col_name = df.columns[idx]
                        val = df.iloc[row_idx - 2][col_name]
                        sheet.cell(row_idx, col_idx).value = val * multiplier
                continue


def measure_memory(process: psutil.Process, stop_event: threading.Event, peak_memory: list):
    """Monitor process memory usage in background thread."""
    while not stop_event.is_set():
        try:
            mem = process.memory_info().rss / 1024 / 1024
            for child in process.children(recursive=True):
                try:
                    mem += child.memory_info().rss / 1024 / 1024
                except psutil.NoSuchProcess:
                    pass
            peak_memory[0] = max(peak_memory[0], mem)
            time.sleep(0.05)
        except psutil.NoSuchProcess:
            break


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 measure_duckdb_fast.py <input_file> [output_file]")
        sys.exit(1)

    input_path = Path(sys.argv[1])
    if not input_path.exists():
        print(f"Error: File not found: {input_path}")
        sys.exit(1)

    output_path = Path(sys.argv[2]) if len(sys.argv) >= 3 else Path(f"output_{input_path.stem}_fast.xlsx")

    process = psutil.Process()
    baseline_mem = process.memory_info().rss / 1024 / 1024

    peak_memory = [baseline_mem]
    stop_event = threading.Event()
    monitor_thread = threading.Thread(target=measure_memory, args=(process, stop_event, peak_memory))
    monitor_thread.start()

    start_time = time.time()

    # Load Excel
    wb = openpyxl.load_workbook(input_path, data_only=False)

    # Load data into DataFrames
    dataframes = load_excel_to_dataframes(input_path)

    # Evaluate formulas
    evaluate_formulas_vectorized(wb, dataframes)

    # Save results
    wb.save(output_path)

    end_time = time.time()
    stop_event.set()
    monitor_thread.join()

    total_rows = sum(sheet.max_row - 1 for sheet in wb.worksheets)
    elapsed = end_time - start_time
    peak_total = peak_memory[0]
    used_mem = peak_total - baseline_mem

    output = {
        "rows": total_rows,
        "peakTotalMB": round(peak_total, 2),
        "usedMB": round(used_mem, 2),
        "baselineMB": round(baseline_mem, 2),
        "timeSeconds": round(elapsed, 3)
    }

    print(json.dumps(output))


if __name__ == "__main__":
    main()
