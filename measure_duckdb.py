#!/usr/bin/env python3
"""
DuckDB Excel Formula Evaluation Benchmark

Two-phase approach:
1. Phase 1: Aggregate functions (SUM, SUMIF, COUNTIF, etc.) → DuckDB SQL
2. Phase 2: Scalar expressions (IF, arithmetic, cell refs) → formulas library
"""

import re
import duckdb
import openpyxl
import pandas as pd
import psutil
import formulas
import json
import sys
import time
import threading
from pathlib import Path
from typing import Any, Dict, List, Tuple, Optional


# ─────────────────────────────────────────────────────────────────
# 1. Formula Parser - Two Phase Evaluation
# ─────────────────────────────────────────────────────────────────

class FormulaEvaluator:
    """Evaluate Excel formulas using DuckDB + formulas library."""

    def __init__(self, conn: duckdb.DuckDBPyConnection, schema: Dict[str, Dict[str, str]]):
        self.conn = conn
        self.schema = schema  # {"sheet1": {"A": "col_a", "B": "col_b"}, ...}

        # Aggregate patterns (most specific first)
        self.aggregate_patterns = [
            (re.compile(r'([A-Za-z0-9_]+)!([A-Z]+)(\d+)', re.I), self._resolve_cross_sheet_ref),
            (re.compile(r'SUM\(([A-Z]):([A-Z])\)', re.I), self._resolve_sum_range),
            (re.compile(r'AVERAGE\(([A-Z]):([A-Z])\)', re.I), self._resolve_average_range),
            (re.compile(r'MAX\(([A-Z]):([A-Z])\)', re.I), self._resolve_max_range),
            (re.compile(r'MIN\(([A-Z]):([A-Z])\)', re.I), self._resolve_min_range),
            (re.compile(r'COUNT\(([A-Z]):([A-Z])\)', re.I), self._resolve_count_range),
        ]

    def _col_name(self, sheet: str, letter: str) -> str:
        """Get DuckDB column name from Excel column letter."""
        return self.schema.get(sheet.lower(), {}).get(letter.upper(), f'col_{letter}')

    def _resolve_cross_sheet_ref(self, m: re.Match) -> str:
        """Resolve cross-sheet reference like Sheet1!A2."""
        sheet_name = m.group(1).lower()
        col_letter = m.group(2).upper()
        row_num = int(m.group(3))
        col_name = self._col_name(sheet_name, col_letter)

        result = self.conn.execute(
            f"SELECT {col_name} FROM {sheet_name} WHERE _row = ?",
            [row_num]
        ).fetchone()

        if result and result[0] is not None:
            return str(float(result[0]))
        return "0"

    def _resolve_sum_range(self, m: re.Match) -> str:
        col = self._col_name("sheet1", m.group(1))
        result = self.conn.execute(f"SELECT COALESCE(SUM({col}), 0) FROM sheet1").fetchone()[0]
        return str(float(result))

    def _resolve_average_range(self, m: re.Match) -> str:
        col = self._col_name("sheet1", m.group(1))
        result = self.conn.execute(f"SELECT COALESCE(AVG({col}), 0) FROM sheet1").fetchone()[0]
        return str(float(result))

    def _resolve_max_range(self, m: re.Match) -> str:
        col = self._col_name("sheet1", m.group(1))
        result = self.conn.execute(f"SELECT COALESCE(MAX({col}), 0) FROM sheet1").fetchone()[0]
        return str(float(result))

    def _resolve_min_range(self, m: re.Match) -> str:
        col = self._col_name("sheet1", m.group(1))
        result = self.conn.execute(f"SELECT COALESCE(MIN({col}), 0) FROM sheet1").fetchone()[0]
        return str(float(result))

    def _resolve_count_range(self, m: re.Match) -> str:
        col = self._col_name("sheet1", m.group(1))
        result = self.conn.execute(f"SELECT COUNT(*) FROM sheet1 WHERE {col} IS NOT NULL").fetchone()[0]
        return str(int(result))

    def resolve_aggregates(self, formula: str) -> str:
        """Phase 1: Replace aggregate functions with scalar values."""
        for _ in range(10):  # Safety cap for nesting
            changed = False
            for pattern, resolver in self.aggregate_patterns:
                new = pattern.sub(resolver, formula)
                if new != formula:
                    formula = new
                    changed = True
            if not changed:
                break
        return formula

    def eval_scalar(self, formula: str, row_ctx: dict) -> float:
        """Phase 2: Evaluate scalar expression with formulas library."""
        expr = formula.lstrip("=").strip()

        # Substitute cell references with values
        for ref, value in row_ctx.items():
            literal = f'"{value}"' if isinstance(value, str) else str(value)
            expr = re.sub(rf'\b{re.escape(ref)}\b', literal, expr, flags=re.I)

        try:
            result = formulas.Parser().ast("=" + expr)[1].compile()()
            return float(result)
        except Exception as e:
            raise ValueError(f"formulas eval failed: '={expr}' → {e}") from e

    def evaluate(self, formula: str, row_ctx: dict) -> float:
        """Evaluate formula: Phase 1 (aggregates) → Phase 2 (scalar)."""
        simplified = self.resolve_aggregates(formula)
        return self.eval_scalar(simplified, row_ctx)


# ─────────────────────────────────────────────────────────────────
# 2. Excel Data Loading
# ─────────────────────────────────────────────────────────────────

def load_excel_to_duckdb(excel_path: Path, conn: duckdb.DuckDBPyConnection) -> Tuple[Dict[str, Dict[str, str]], openpyxl.Workbook]:
    """Load Excel sheets into DuckDB tables. Returns (schema, workbook)."""
    wb = openpyxl.load_workbook(excel_path, read_only=False, data_only=False)
    schema = {}

    for sheet in wb.worksheets:
        sheet_name = sheet.title.replace(' ', '_').lower()

        # Read all data into list of dicts
        data = []
        max_col = sheet.max_column

        # Get headers from row 1
        headers = []
        for col_idx in range(1, max_col + 1):
            cell_val = sheet.cell(1, col_idx).value
            col_name = str(cell_val).lower().replace(' ', '_') if cell_val else f'col_{col_idx}'
            headers.append(col_name)

        # Build column letter mapping
        col_mapping = {}
        for col_idx, header in enumerate(headers):
            col_letter = chr(65 + col_idx) if col_idx < 26 else f'{{chr(65 + (col_idx // 26 - 1))}}{chr(65 + (col_idx % 26))}'
            col_letter = chr(65 + col_idx)  # Simple A, B, C...
            col_mapping[col_letter.upper()] = header

        schema[sheet_name] = col_mapping

        # Read data rows (starting from row 2)
        for row_idx in range(2, sheet.max_row + 1):
            row_dict = {'_row': row_idx}
            for col_idx, header in enumerate(headers):
                cell_val = sheet.cell(row_idx, col_idx + 1).value
                row_dict[header] = cell_val
            data.append(row_dict)

        if data:
            # Create DataFrame and register with DuckDB
            df = pd.DataFrame(data)
            conn.register(f'{sheet_name}_df', df)
            conn.execute(f'CREATE TABLE {sheet_name} AS SELECT * FROM {sheet_name}_df')

    return schema, wb


# ─────────────────────────────────────────────────────────────────
# 3. Formula Evaluation
# ─────────────────────────────────────────────────────────────────

def evaluate_all_formulas(
    conn: duckdb.DuckDBPyConnection,
    wb: openpyxl.Workbook,
    evaluator: FormulaEvaluator
) -> List[Tuple[str, str, float]]:
    """Evaluate all formulas in workbook. Returns list of (cell, formula, result)."""
    results = []

    for sheet in wb.worksheets:
        sheet_name = sheet.title

        for row_idx in range(2, sheet.max_row + 1):
            # Build row context for this row
            row_ctx = {}
            for col_idx in range(1, sheet.max_column + 1):
                cell = sheet.cell(row_idx, col_idx)
                col_letter = chr(64 + col_idx)  # A, B, C...
                if cell.value is not None and not cell.data_type == 'f':
                    row_ctx[f'{col_letter}{row_idx}'] = cell.value

            # Evaluate formula cells
            for col_idx in range(1, sheet.max_column + 1):
                cell = sheet.cell(row_idx, col_idx)
                if cell.data_type == 'f':  # Formula cell
                    try:
                        result = evaluator.evaluate(cell.value, row_ctx)
                        results.append((f'{sheet.title}!{cell.coordinate}', cell.value, result))
                        cell.value = result  # Write result back
                    except Exception as e:
                        results.append((f'{sheet.title}!{cell.coordinate}', cell.value, str(e)))

    return results


# ─────────────────────────────────────────────────────────────────
# 4. Memory & Timing Measurement
# ─────────────────────────────────────────────────────────────────

def measure_memory(process: psutil.Process, stop_event: threading.Event, peak_memory: list):
    """Monitor process memory usage in background thread."""
    while not stop_event.is_set():
        try:
            mem = process.memory_info().rss / 1024 / 1024
            for child in process.children(recursive=True):
                try:
                    mem += child.memory_info().rss / 1024 / 1024
                except psutil.NoSuchProcess:
                    pass
            peak_memory[0] = max(peak_memory[0], mem)
            time.sleep(0.05)
        except psutil.NoSuchProcess:
            break


# ─────────────────────────────────────────────────────────────────
# 5. Main
# ─────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("Usage: python3 measure_duckdb.py <input_file> [output_file]")
        sys.exit(1)

    input_path = Path(sys.argv[1])
    if not input_path.exists():
        print(f"Error: File not found: {input_path}")
        sys.exit(1)

    output_path = Path(sys.argv[2]) if len(sys.argv) >= 3 else Path(f"output_{input_path.stem}_duckdb.xlsx")

    process = psutil.Process()
    baseline_mem = process.memory_info().rss / 1024 / 1024

    # Start memory monitoring
    peak_memory = [baseline_mem]
    stop_event = threading.Event()
    monitor_thread = threading.Thread(target=measure_memory, args=(process, stop_event, peak_memory))
    monitor_thread.start()

    start_time = time.time()

    # Load Excel into DuckDB
    conn = duckdb.connect(':memory:')
    schema, wb = load_excel_to_duckdb(input_path, conn)

    # Evaluate formulas
    evaluator = FormulaEvaluator(conn, schema)
    evaluate_all_formulas(conn, wb, evaluator)

    # Save results
    wb.save(output_path)

    end_time = time.time()

    # Stop monitoring
    stop_event.set()
    monitor_thread.join()

    # Calculate metrics
    total_rows = sum(sheet.max_row - 1 for sheet in wb.worksheets)
    elapsed = end_time - start_time
    peak_total = peak_memory[0]
    used_mem = peak_total - baseline_mem

    output = {
        "rows": total_rows,
        "peakTotalMB": round(peak_total, 2),
        "usedMB": round(used_mem, 2),
        "baselineMB": round(baseline_mem, 2),
        "timeSeconds": round(elapsed, 3)
    }

    print(json.dumps(output))
    conn.close()


if __name__ == "__main__":
    main()
